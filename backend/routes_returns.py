"""Return master & query management routes."""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import Response
from typing import Optional
from datetime import datetime, timezone, date
import io
import re
from pymongo.errors import DuplicateKeyError
from models import (
    ReturnCreate, ReturnUpdate, ReassignRequest,
    QueryCreate, QueryUpdate, gen_id, utcnow
)
from auth import get_current_user, require_admin
from db import get_db
from audit import log_audit
from utils_export import to_csv, to_xlsx, to_pdf, read_csv_bytes, read_xlsx_bytes
from escalation import compute_breaches, return_stage_age_days, return_total_age_days

router = APIRouter(tags=["operations"])

RETURN_IMPORT_ALIASES = {
    "returninwardno": "return_inward_no",
    "itrinwardno": "return_inward_no",
    "inwardno": "return_inward_no",
    "returninwarddate": "return_inward_date",
    "inwarddate": "return_inward_date",
    "date": "return_inward_date",
    "taskiditr": "task_id",
    "taskid": "task_id",
    "fy": "fy",
    "financialyear": "fy",
    "fileno": "file_no",
    "group": "group",
    "clientnames": "client_name",
    "clientname": "client_name",
    "returntype": "return_type",
    "itrform": "itr_form",
    "itrformfy2425": "itr_form_previous",
    "itrformfy202425": "itr_form_previous",
    "itrformfy202526": "itr_form",
    "duedate": "due_date",
    "itrstatus": "current_stage_name",
    "currentstage": "current_stage_name",
    "currentstagename": "current_stage_name",
    "personassigned": "person_assigned",
    "personassignedname": "person_assigned_name",
    "personassignedemail": "person_assigned_email",
    "remarks": "remarks",
}


def _normalise_return_header(value) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def _normalise_return_lookup(value) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def _clean_return_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\t", " ").strip()


def _clean_return_date(value) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _clean_return_text(value)
    if not text:
        return ""
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return text


def _clean_itr_form(value) -> str:
    text = _clean_return_text(value)
    if not text:
        return ""
    if re.fullmatch(r"\d+(\.0)?", text):
        return f"ITR-{int(float(text))}"
    match = re.fullmatch(r"itr[-\s]*(\d+)", text, flags=re.IGNORECASE)
    if match:
        return f"ITR-{match.group(1)}"
    return text


def _looks_like_return_import_header(headers: list) -> bool:
    keys = {_normalise_return_header(h) for h in headers}
    has_native_return_headers = "returninwardno" in keys and "clientname" in keys
    has_report_return_headers = "itrinwardno" in keys and "clientnames" in keys and "itrstatus" in keys
    return has_native_return_headers or has_report_return_headers


def _map_return_import_row(row: dict) -> dict:
    mapped = {}
    for raw_key, value in row.items():
        canonical = RETURN_IMPORT_ALIASES.get(_normalise_return_header(raw_key))
        if not canonical:
            continue
        if canonical == "itr_form_previous":
            if not mapped.get("itr_form"):
                mapped["itr_form"] = _clean_itr_form(value)
            continue
        if canonical in ("return_inward_date", "due_date"):
            mapped[canonical] = _clean_return_date(value)
        elif canonical == "itr_form":
            itr_form = _clean_itr_form(value)
            if itr_form or not mapped.get(canonical):
                mapped[canonical] = itr_form
        else:
            mapped[canonical] = _clean_return_text(value)
    return mapped


def _read_return_import_rows(content: bytes, filename: str) -> list[dict]:
    name = filename.lower()
    if name.endswith(".csv"):
        return [_map_return_import_row(row) for row in read_csv_bytes(content)]
    if not name.endswith(".xlsx"):
        raise HTTPException(400, "Unsupported file type. Use .csv or .xlsx")

    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    best_rows = []
    for ws in wb.worksheets:
        candidate_rows = []
        raw_rows = ws.iter_rows(values_only=True)
        for header_row in raw_rows:
            headers = list(header_row or [])
            if not _looks_like_return_import_header(headers):
                continue
            for values in raw_rows:
                if all(v is None or v == "" for v in values):
                    continue
                row = {str(headers[i] or ""): (values[i] if i < len(values) else "") for i in range(len(headers))}
                mapped = _map_return_import_row(row)
                if mapped.get("return_inward_no") and mapped.get("client_name"):
                    candidate_rows.append(mapped)
            if len(candidate_rows) > len(best_rows):
                best_rows = candidate_rows
            break
    if best_rows:
        return best_rows

    return [_map_return_import_row(row) for row in read_xlsx_bytes(content)]


def _is_import_status_query(status: str) -> bool:
    normalised = _normalise_return_lookup(status)
    return bool(normalised) and (
        "onlycgdatareceived" in normalised
        or "pendingentry" in normalised
        or "entrydone" in normalised
    )


def _query_text_from_import_status(status: str) -> str:
    return re.sub(r"\s*-\s*", " - ", _clean_return_text(status)).strip()


def _date_as_stage_timestamp(value: str, fallback: str) -> str:
    cleaned = _clean_return_date(value)
    if not cleaned:
        return fallback
    try:
        return datetime.fromisoformat(cleaned).replace(tzinfo=timezone.utc).isoformat()
    except ValueError:
        return fallback


def _is_initial_return_stage(stage: Optional[dict]) -> bool:
    if not stage:
        return False
    return int(stage.get("sequence") or 0) == 1 or _normalise_return_lookup(stage.get("stage_name")) == "documentinward"


async def _enrich_return(db, r, stage_map=None, user_map=None):
    if stage_map is None:
        stages = await db.workflow_stages.find({}, {"_id": 0}).to_list(200)
        stage_map = {s["id"]: s for s in stages}
    if user_map is None:
        users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(500)
        user_map = {u["id"]: u for u in users}
    stage = stage_map.get(r.get("current_stage_id"))
    assignee = user_map.get(r.get("person_assigned_id")) if r.get("person_assigned_id") else None
    return {
        **{k: v for k, v in r.items() if k != "_id"},
        "current_stage_name": stage.get("stage_name") if stage else None,
        "current_stage_colour": stage.get("dashboard_colour") if stage else None,
        "next_action_required": stage.get("next_action_required") if stage else "",
        "person_assigned_name": assignee.get("name") if assignee else None,
        "stage_ageing_days": return_stage_age_days(r),
        "total_ageing_days": return_total_age_days(r),
        "last_updated_date": r.get("updated_at"),
    }


@router.get("/returns")
async def list_returns(
    search: Optional[str] = None,
    stage_id: Optional[str] = None,
    person_id: Optional[str] = None,
    return_type: Optional[str] = None,
    fy: Optional[str] = None,
    overdue: Optional[bool] = None,
    dashboard_filter: Optional[str] = None,
    stage_age_bucket: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    db = get_db()
    stages = await db.workflow_stages.find({}, {"_id": 0}).to_list(200)
    stage_map = {s["id"]: s for s in stages}
    stage_name_by_id = {s["id"]: s.get("stage_name") or "" for s in stages}
    q = {}
    if dashboard_filter and dashboard_filter not in {"total_returns", "overdue_returns", "sla_breaches", "upcoming_sla", "escalation_breaches"}:
        filter_stage_ids = []
        if dashboard_filter == "pending_verification":
            keywords = {"documentinward", "dvp", "dvq", "dvqt", "dvnq"}
            filter_stage_ids = [
                s["id"] for s in stages
                if _normalise_return_lookup(s.get("stage_name")) in keywords
            ]
        elif dashboard_filter == "ready_to_file":
            filter_stage_ids = [s["id"] for s in stages if (s.get("stage_name") or "").upper() == "RTG"]
        elif dashboard_filter == "everification_pending":
            filter_stage_ids = [s["id"] for s in stages if (s.get("stage_name") or "") == "Filed-Everification Pending"]
        elif dashboard_filter == "completed_returns":
            filter_stage_ids = [s["id"] for s in stages if (s.get("stage_name") or "").lower() == "completed"]
        if filter_stage_ids:
            q["current_stage_id"] = {"$in": filter_stage_ids}
    if stage_id:
        q["current_stage_id"] = stage_id
    if person_id:
        q["person_assigned_id"] = person_id
    if return_type:
        q["return_type"] = return_type
    if fy:
        q["fy"] = fy
    if search:
        q["$or"] = [
            {"return_inward_no": {"$regex": search, "$options": "i"}},
            {"client_name": {"$regex": search, "$options": "i"}},
            {"file_no": {"$regex": search, "$options": "i"}},
            {"task_id": {"$regex": search, "$options": "i"}},
        ]
    docs = await db.returns.find(q, {"_id": 0}).sort("created_at", -1).to_list(5000)
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(500)
    user_map = {u["id"]: u for u in users}
    enriched = [await _enrich_return(db, r, stage_map, user_map) for r in docs]
    if overdue or dashboard_filter == "overdue_returns":
        today = datetime.now(timezone.utc).date().isoformat()
        enriched = [e for e in enriched if e.get("due_date") and e["due_date"] < today and (e.get("current_stage_name") or "").lower() != "completed"]
    if dashboard_filter in {"sla_breaches", "upcoming_sla", "escalation_breaches"}:
        breaches = await compute_breaches(db)
        key = {
            "sla_breaches": "sla_breaches",
            "upcoming_sla": "upcoming_sla_breaches",
            "escalation_breaches": "escalation_breaches",
        }[dashboard_filter]
        allowed_ids = {x["return_id"] for x in breaches[key]}
        enriched = [e for e in enriched if e.get("id") in allowed_ids]
    if stage_age_bucket:
        def in_bucket(days: int) -> bool:
            if stage_age_bucket == "0-3":
                return days <= 3
            if stage_age_bucket == "4-7":
                return 4 <= days <= 7
            if stage_age_bucket == "8-15":
                return 8 <= days <= 15
            if stage_age_bucket == "15+":
                return days > 15
            return True
        enriched = [
            e for e in enriched
            if (stage_name_by_id.get(e.get("current_stage_id")) or "").lower() != "completed"
            and in_bucket(int(e.get("stage_ageing_days") or 0))
        ]
    return enriched


@router.get("/returns/{rid}")
async def get_return(rid: str, user: dict = Depends(get_current_user)):
    db = get_db()
    r = await db.returns.find_one({"id": rid}, {"_id": 0})
    if not r:
        raise HTTPException(404, "Not found")
    return await _enrich_return(db, r)


@router.post("/returns")
async def create_return(payload: ReturnCreate, admin: dict = Depends(require_admin)):
    db = get_db()
    stage = await db.workflow_stages.find_one({"id": payload.current_stage_id})
    if not stage:
        raise HTTPException(400, "Invalid stage")
    now = utcnow().isoformat()
    stage_entered_at = _date_as_stage_timestamp(payload.return_inward_date, now) if _is_initial_return_stage(stage) else now
    doc = {
        "id": gen_id(),
        **payload.model_dump(),
        "stage_entered_at": stage_entered_at,
        "created_at": now,
        "updated_at": now,
    }
    try:
        await db.returns.insert_one(doc)
    except DuplicateKeyError:
        raise HTTPException(status_code=409, detail=f"Return with inward no. '{payload.return_inward_no}' already exists")
    await log_audit(db, admin, "Returns", "Create", entity_id=doc["id"], new_value={"return_inward_no": payload.return_inward_no, "stage": stage["stage_name"]})
    doc.pop("_id", None)
    return await _enrich_return(db, doc)


@router.patch("/returns/{rid}")
async def update_return(rid: str, payload: ReturnUpdate, admin: dict = Depends(require_admin)):
    db = get_db()
    existing = await db.returns.find_one({"id": rid})
    if not existing:
        raise HTTPException(404, "Not found")
    update = payload.model_dump(exclude_none=True)
    if "current_stage_id" in update and update["current_stage_id"] != existing.get("current_stage_id"):
        update["stage_entered_at"] = utcnow().isoformat()
        old_stage = await db.workflow_stages.find_one({"id": existing.get("current_stage_id")})
        new_stage = await db.workflow_stages.find_one({"id": update["current_stage_id"]})
        await log_audit(db, admin, "Returns", "Stage Changed", entity_id=rid,
                        old_value=old_stage["stage_name"] if old_stage else None,
                        new_value=new_stage["stage_name"] if new_stage else None)
    elif "return_inward_date" in update and update["return_inward_date"] != existing.get("return_inward_date"):
        current_stage = await db.workflow_stages.find_one({"id": existing.get("current_stage_id")})
        if _is_initial_return_stage(current_stage):
            update["stage_entered_at"] = _date_as_stage_timestamp(update["return_inward_date"], existing.get("stage_entered_at") or utcnow().isoformat())
    if "person_assigned_id" in update and update["person_assigned_id"] != existing.get("person_assigned_id"):
        old_user = await db.users.find_one({"id": existing.get("person_assigned_id")}) if existing.get("person_assigned_id") else None
        new_user = await db.users.find_one({"id": update["person_assigned_id"]}) if update["person_assigned_id"] else None
        await log_audit(db, admin, "Returns", "Person Assigned Changed", entity_id=rid,
                        old_value=old_user["name"] if old_user else None,
                        new_value=new_user["name"] if new_user else None)
    update["updated_at"] = utcnow().isoformat()
    await db.returns.update_one({"id": rid}, {"$set": update})
    return {"ok": True}


@router.post("/returns/{rid}/reassign")
async def reassign_return(rid: str, payload: ReassignRequest, admin: dict = Depends(require_admin)):
    db = get_db()
    existing = await db.returns.find_one({"id": rid})
    if not existing:
        raise HTTPException(404, "Not found")
    old_user = await db.users.find_one({"id": existing.get("person_assigned_id")}) if existing.get("person_assigned_id") else None
    new_user = await db.users.find_one({"id": payload.person_assigned_id}) if payload.person_assigned_id else None
    await db.returns.update_one({"id": rid}, {"$set": {"person_assigned_id": payload.person_assigned_id, "updated_at": utcnow().isoformat()}})
    await log_audit(db, admin, "Returns", "Person Assigned Changed", entity_id=rid,
                    old_value=old_user["name"] if old_user else None,
                    new_value=new_user["name"] if new_user else None)
    return {"ok": True}


@router.delete("/returns/{rid}")
async def delete_return(rid: str, admin: dict = Depends(require_admin)):
    db = get_db()
    await db.returns.delete_one({"id": rid})
    await db.queries.delete_many({"return_id": rid})
    await log_audit(db, admin, "Returns", "Delete", entity_id=rid)
    return {"ok": True}


RETURN_EXPORT_HEADERS = [
    "return_inward_no", "return_inward_date", "task_id", "fy", "file_no", "group",
    "client_name", "return_type", "itr_form", "due_date",
    "current_stage_name", "next_action_required", "person_assigned_name",
    "stage_ageing_days", "total_ageing_days", "last_updated_date"
]


@router.get("/returns/export/file")
async def export_returns(format: str = Query("xlsx"), user: dict = Depends(get_current_user)):
    db = get_db()
    docs = await db.returns.find({}, {"_id": 0}).sort("created_at", -1).to_list(10000)
    stages = await db.workflow_stages.find({}, {"_id": 0}).to_list(200)
    stage_map = {s["id"]: s for s in stages}
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(500)
    user_map = {u["id"]: u for u in users}
    enriched = [await _enrich_return(db, r, stage_map, user_map) for r in docs]
    if format == "csv":
        data = to_csv(enriched, RETURN_EXPORT_HEADERS)
        return Response(content=data, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=returns.csv"})
    if format == "pdf":
        data = to_pdf(enriched, RETURN_EXPORT_HEADERS, "Return Master Report")
        return Response(content=data, media_type="application/pdf", headers={"Content-Disposition": "attachment; filename=returns.pdf"})
    data = to_xlsx(enriched, RETURN_EXPORT_HEADERS, "Returns")
    return Response(content=data, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=returns.xlsx"})


@router.post("/returns/import")
async def import_returns(file: UploadFile = File(...), admin: dict = Depends(require_admin)):
    db = get_db()
    content = await file.read()
    rows = _read_return_import_rows(content, file.filename or "")
    stages = await db.workflow_stages.find({}, {"_id": 0}).to_list(200)
    stage_by_name = {s["stage_name"].lower(): s for s in stages}
    stage_by_lookup = {_normalise_return_lookup(s["stage_name"]): s for s in stages}
    default_stage = (
        stage_by_lookup.get("documentinward")
        or stage_by_name.get("document inward")
        or stages[0]
    )
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(500)
    user_by_email = {u["email"].lower(): u for u in users}
    user_by_name = {u["name"].lower(): u for u in users}

    inserted = 0
    skipped = 0
    for row in rows:
        rin = str(row.get("return_inward_no") or "").strip()
        if not rin:
            skipped += 1
            continue
        now = utcnow().isoformat()
        raw_stage_name = str(row.get("current_stage_name") or row.get("current_stage") or "").strip()
        stage_name = raw_stage_name.lower()
        import_status_is_query = _is_import_status_query(raw_stage_name)
        stage = None
        if not import_status_is_query:
            stage = stage_by_name.get(stage_name) or stage_by_lookup.get(_normalise_return_lookup(raw_stage_name))
        if not stage:
            stage = default_stage
        assignee_field = str(row.get("person_assigned") or row.get("person_assigned_name") or row.get("person_assigned_email") or "").strip().lower()
        assignee = user_by_email.get(assignee_field) or user_by_name.get(assignee_field)
        doc = {
            "id": gen_id(),
            "return_inward_no": rin,
            "return_inward_date": str(row.get("return_inward_date") or "").strip(),
            "task_id": str(row.get("task_id") or "").strip(),
            "fy": str(row.get("fy") or "2025-26").strip(),
            "file_no": str(row.get("file_no") or "").strip(),
            "group": str(row.get("group") or "").strip(),
            "client_name": str(row.get("client_name") or "").strip(),
            "return_type": str(row.get("return_type") or "Original").strip(),
            "itr_form": str(row.get("itr_form") or "").strip(),
            "due_date": str(row.get("due_date") or "").strip(),
            "current_stage_id": stage["id"],
            "person_assigned_id": assignee["id"] if assignee else None,
            "remarks": str(row.get("remarks") or "").strip(),
            "stage_entered_at": _date_as_stage_timestamp(row.get("return_inward_date"), now),
            "created_at": now,
            "updated_at": now,
        }
        if await db.returns.find_one({"return_inward_no": rin}):
            skipped += 1
            continue
        await db.returns.insert_one(doc)
        if import_status_is_query:
            query_doc = {
                "id": gen_id(),
                "return_id": doc["id"],
                "query_raised_by_id": admin["id"],
                "query_raised_date": utcnow().date().isoformat(),
                "query_description": _query_text_from_import_status(raw_stage_name),
                "query_status": "Open",
                "follow_up_date": None,
                "query_closed_date": None,
                "remarks": "Created from imported ITR Status",
                "created_at": now,
                "updated_at": now,
            }
            await db.queries.insert_one(query_doc)
        inserted += 1
    await log_audit(db, admin, "Returns", "Import", new_value={"inserted": inserted, "skipped": skipped})
    return {"inserted": inserted, "skipped": skipped}


# =================== Queries ===================
async def _enrich_query(db, q, user_map=None, return_map=None):
    if user_map is None:
        users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(500)
        user_map = {u["id"]: u for u in users}
    if return_map is None:
        ret = await db.returns.find_one({"id": q.get("return_id")}, {"_id": 0}) if q.get("return_id") else None
    else:
        ret = return_map.get(q.get("return_id"))
    raised_by = user_map.get(q.get("query_raised_by_id")) if q.get("query_raised_by_id") else None
    return {
        **{k: v for k, v in q.items() if k != "_id"},
        "return_inward_no": ret.get("return_inward_no") if ret else None,
        "client_name": ret.get("client_name") if ret else None,
        "file_no": ret.get("file_no") if ret else None,
        "group": ret.get("group") if ret else None,
        "fy": ret.get("fy") if ret else None,
        "query_raised_by_name": raised_by.get("name") if raised_by else None,
    }


@router.get("/queries")
async def list_queries(
    return_id: Optional[str] = None,
    status: Optional[str] = None,
    pending: Optional[bool] = None,
    search: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    db = get_db()
    q = {}
    if return_id:
        q["return_id"] = return_id
    if pending:
        q["query_status"] = {"$ne": "Closed"}
    elif status:
        q["query_status"] = status
    if search:
        q["query_description"] = {"$regex": search, "$options": "i"}
    docs = await db.queries.find(q, {"_id": 0}).sort("created_at", -1).to_list(5000)
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(500)
    user_map = {u["id"]: u for u in users}
    return_ids = list({x.get("return_id") for x in docs if x.get("return_id")})
    returns = await db.returns.find({"id": {"$in": return_ids}}, {"_id": 0}).to_list(len(return_ids)) if return_ids else []
    return_map = {r["id"]: r for r in returns}
    return [await _enrich_query(db, x, user_map, return_map) for x in docs]


@router.post("/queries")
async def create_query(payload: QueryCreate, user: dict = Depends(get_current_user)):
    db = get_db()
    ret = await db.returns.find_one({"id": payload.return_id})
    if not ret:
        raise HTTPException(400, "Invalid return")
    now = utcnow().isoformat()
    doc = {
        "id": gen_id(),
        **payload.model_dump(),
        "query_raised_by_id": payload.query_raised_by_id or user["id"],
        "query_raised_date": payload.query_raised_date or utcnow().date().isoformat(),
        "query_closed_date": None,
        "created_at": now,
        "updated_at": now,
    }
    await db.queries.insert_one(doc)
    await log_audit(db, user, "Queries", "Query Added", entity_id=doc["id"], new_value={"return_id": payload.return_id, "description": payload.query_description[:80]})
    doc.pop("_id", None)
    return await _enrich_query(db, doc)


def _normalise_query_value(key: str, value):
    if key in {"query_description", "remarks", "follow_up_date", "query_closed_date"} and value is None:
        return ""
    return value


@router.patch("/queries/{qid}")
async def update_query(qid: str, payload: QueryUpdate, user: dict = Depends(get_current_user)):
    db = get_db()
    existing = await db.queries.find_one({"id": qid})
    if not existing:
        raise HTTPException(404, "Query not found")
    update = payload.model_dump(exclude_unset=True)
    # auto-close
    if update.get("query_status") == "Closed" and not existing.get("query_closed_date"):
        update["query_closed_date"] = utcnow().date().isoformat()
    update = {
        k: v
        for k, v in update.items()
        if _normalise_query_value(k, existing.get(k)) != _normalise_query_value(k, v)
    }
    if not update:
        return {"ok": True, "changed": False}
    update["updated_at"] = utcnow().isoformat()
    await db.queries.update_one({"id": qid}, {"$set": update})
    action = "Query Closed" if update.get("query_status") == "Closed" else "Query Updated"
    await log_audit(db, user, "Queries", action, entity_id=qid,
                    old_value={k: existing.get(k) for k in update},
                    new_value=update)
    return {"ok": True}


@router.delete("/queries/{qid}")
async def delete_query(qid: str, admin: dict = Depends(require_admin)):
    db = get_db()
    await db.queries.delete_one({"id": qid})
    await log_audit(db, admin, "Queries", "Delete", entity_id=qid)
    return {"ok": True}


QUERY_EXPORT_HEADERS = [
    "id", "return_id", "return_inward_no", "client_name", "file_no", "group", "fy",
    "query_raised_by_name", "query_raised_date",
    "query_description", "query_status", "follow_up_date", "query_closed_date", "remarks"
]


@router.get("/queries/export/file")
async def export_queries(format: str = Query("xlsx"), status: Optional[str] = None, pending: Optional[bool] = None, user: dict = Depends(get_current_user)):
    db = get_db()
    q = {}
    if pending:
        q["query_status"] = {"$ne": "Closed"}
    elif status:
        q["query_status"] = status
    docs = await db.queries.find(q, {"_id": 0}).sort("created_at", -1).to_list(10000)
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(500)
    user_map = {u["id"]: u for u in users}
    return_ids = list({x.get("return_id") for x in docs if x.get("return_id")})
    returns = await db.returns.find({"id": {"$in": return_ids}}, {"_id": 0}).to_list(len(return_ids)) if return_ids else []
    return_map = {r["id"]: r for r in returns}
    enriched = [await _enrich_query(db, x, user_map, return_map) for x in docs]
    if format == "csv":
        data = to_csv(enriched, QUERY_EXPORT_HEADERS)
        return Response(content=data, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=queries.csv"})
    if format == "pdf":
        data = to_pdf(enriched, QUERY_EXPORT_HEADERS, "Queries Report")
        return Response(content=data, media_type="application/pdf", headers={"Content-Disposition": "attachment; filename=queries.pdf"})
    data = to_xlsx(enriched, QUERY_EXPORT_HEADERS, "Queries")
    return Response(content=data, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=queries.xlsx"})
