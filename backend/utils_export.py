"""Export helpers for Excel, CSV, PDF."""
import io
import csv
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet


def to_csv(rows: list[dict], headers: list[str]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for r in rows:
        writer.writerow({h: r.get(h, "") for h in headers})
    return buf.getvalue().encode("utf-8")


def to_xlsx(rows: list[dict], headers: list[str], sheet_name: str = "Export") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    for i, h in enumerate(headers, 1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = max(12, min(40, len(h) + 4))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_pdf(rows: list[dict], headers: list[str], title: str = "Report") -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=20, bottomMargin=20, leftMargin=20, rightMargin=20)
    styles = getSampleStyleSheet()
    story = [Paragraph(title, styles["Title"]), Spacer(1, 10)]
    data = [headers]
    for r in rows:
        data.append([str(r.get(h, "") or "") for h in headers])
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#065f46")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(table)
    doc.build(story)
    return buf.getvalue()


def read_csv_bytes(data: bytes) -> list[dict]:
    text = data.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def read_xlsx_bytes(data: bytes) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    result = []
    for r in rows[1:]:
        if all(v is None or v == "" for v in r):
            continue
        result.append({headers[i]: ("" if r[i] is None else r[i]) for i in range(len(headers))})
    return result
